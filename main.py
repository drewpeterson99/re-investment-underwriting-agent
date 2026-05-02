import argparse
import json
import math
import os
import re
import shutil
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, Tuple

import yaml
from llama_cpp import Llama
from openpyxl import load_workbook

# strptime formats accepted for ListingDate (stored as m/d/yyyy)
_LISTING_DATE_INPUT_FORMATS = ("%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%Y-%m-%d")


def normalize_extraction_source_text(text: str) -> str:
    """Light cleanup before extraction: BOM, newlines, common PDF fullwidth dollar (＄ → $)."""
    if not text:
        return text
    t = text.replace("\ufeff", "").replace("\r\n", "\n").replace("\r", "\n")
    t = t.replace("\uff04", "$")
    return t.strip()


def read_prompt_via_gui(
    title: str = "Paste listing text",
    initial: str = "",
) -> str | None:
    """
    Multiline GUI prompt so pasted text keeps $ amounts (no shell interpolation).

    Returns None if the user closes the window without confirming usable text.
    Uses tkinter (stdlib). Lazy-import so headless environments without Tk still run CLI-only.
    """
    import tkinter as tk
    from tkinter import scrolledtext

    out: list[str | None] = [None]
    root = tk.Tk()
    root.title(title)
    root.geometry("720x400")
    root.minsize(520, 280)
    root.grid_columnconfigure(0, weight=1)
    root.grid_rowconfigure(1, weight=1)

    hint = tk.Label(
        root,
        text="Paste the listing text below, then click Finish.",
        anchor="w",
        font=("Segoe UI", 10),
    )
    hint.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 4))

    body = scrolledtext.ScrolledText(root, wrap=tk.WORD, font=("Segoe UI", 10), undo=True)
    body.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 8))
    if initial:
        body.insert("1.0", initial)

    row = tk.Frame(root)
    row.grid(row=2, column=0, sticky="ew", padx=10, pady=(0, 12))

    def submit() -> None:
        text = body.get("1.0", tk.END).strip()
        out[0] = text if text else None
        root.destroy()

    def cancel() -> None:
        out[0] = None
        root.destroy()

    tk.Button(row, text="Cancel", width=12, command=cancel).pack(side=tk.RIGHT, padx=(8, 0))
    tk.Button(row, text="Finish", width=12, command=submit, default=tk.ACTIVE).pack(side=tk.RIGHT)
    root.protocol("WM_DELETE_WINDOW", cancel)
    root.mainloop()
    return out[0]


def load_schema(schema_path: Path) -> Dict[str, Any]:
    with schema_path.open("r", encoding="utf-8") as f:
        schema = yaml.safe_load(f)
    if not isinstance(schema, dict) or "fields" not in schema:
        raise ValueError("Schema must be a YAML mapping with a top-level 'fields' key.")
    if not isinstance(schema["fields"], dict) or not schema["fields"]:
        raise ValueError("'fields' must be a non-empty mapping in schema.")
    return schema


def schema_to_instruction_block(schema: Dict[str, Any]) -> str:
    lines = []
    for field_name, field_spec in schema["fields"].items():
        field_type = field_spec.get("type", "string")
        required = field_spec.get("required", False)
        description = field_spec.get("description", "")
        bounds = []
        if "min" in field_spec:
            bounds.append(f"min={field_spec['min']}")
        if "max" in field_spec:
            bounds.append(f"max={field_spec['max']}")
        bounds_text = f" ({', '.join(bounds)})" if bounds else ""
        req_text = "required" if required else "optional"
        normalization = field_spec.get("normalization", {})
        norm_text = ""
        if isinstance(normalization, dict) and normalization.get("rule"):
            norm_text = f" Normalization: {normalization['rule']}."
        fmt = field_spec.get("format")
        fmt_text = f" Format: {fmt}." if fmt else ""
        dp = field_spec.get("decimal_places")
        dp_text = f" At most {dp} decimal place(s)." if dp is not None else ""
        lines.append(
            f"- {field_name}: {field_type}{bounds_text}, {req_text}. {description}{fmt_text}{dp_text}{norm_text}"
        )
    return "\n".join(lines)


def normalize_currency_text_to_int(raw_value: Any) -> int:
    """Parse dollar amounts; supports k/M suffixes (e.g. 675k → 675000) so LLM shorthand parses correctly."""
    text = str(raw_value).strip()
    if not text:
        raise ValueError(f"Could not parse integer value from input: {raw_value!r}")

    t = text.lower().replace(",", "")
    if t.startswith("$"):
        t = t[1:].strip()
    km_m = re.fullmatch(r"(\d+(?:\.\d+)?)\s*([km])\s*", t)
    if km_m:
        val = float(km_m.group(1))
        if km_m.group(2) == "k":
            val *= 1000
        else:
            val *= 1_000_000
        return int(math.ceil(val))

    # Remove currency symbols/codes/punctuation and keep numeric characters.
    text = re.sub(r"[^0-9.\-]", "", t)
    if not text:
        raise ValueError(f"Could not parse integer value from input: {raw_value!r}")
    return int(math.ceil(float(text)))


def _strip_to_numeric_chars(raw_value: Any) -> str:
    """Keep digits, dot, minus—shared by integer and decimal field coercion."""
    return re.sub(r"[^0-9.\-]", "", str(raw_value).strip())


def parse_integer_text(raw_value: Any) -> int:
    """Parse counts and non-currency integers (commas, optional decimals rounded to nearest whole)."""
    text = _strip_to_numeric_chars(raw_value)
    if not text:
        raise ValueError(f"Could not parse integer value from input: {raw_value!r}")
    return int(round(float(text)))


def parse_number_text(raw_value: Any, decimal_places: int) -> float:
    """Parse a numeric string and round to a fixed number of decimal places."""
    text = _strip_to_numeric_chars(raw_value)
    if not text:
        raise ValueError(f"Could not parse number from input: {raw_value!r}")
    return round(float(text), decimal_places)


def normalization_rule_is_currency(field_spec: Dict[str, Any]) -> bool:
    normalization = field_spec.get("normalization", {})
    if not isinstance(normalization, dict):
        return False
    rule = normalization.get("rule", "")
    return isinstance(rule, str) and "currency" in rule.lower()


def normalize_listing_date_m_d_yyyy(raw_value: Any) -> str:
    """Parse a date string and return m/d/yyyy (no leading zeros on month or day)."""
    s = str(raw_value).strip()
    if not s:
        raise ValueError("ListingDate is empty.")
    for fmt in _LISTING_DATE_INPUT_FORMATS:
        try:
            d = datetime.strptime(s, fmt).date()
            return f"{d.month}/{d.day}/{d.year}"
        except ValueError:
            continue
    raise ValueError(
        f"Could not parse date (expected a calendar date, e.g. m/d/yyyy): {raw_value!r}"
    )


def normalize_city_state_string(raw_value: Any) -> str:
    """Normalize to 'City, ST' with a single comma; validate both sides are non-empty."""
    s = " ".join(str(raw_value).split())
    if not s:
        raise ValueError("CityState is empty.")
    if "," not in s:
        raise ValueError(f"CityState must look like 'City, State'. Got: {raw_value!r}")
    left, right = s.split(",", 1)
    city, state = left.strip(), right.strip()
    if not city or not state:
        raise ValueError(f"CityState must include both city and state. Got: {raw_value!r}")
    return f"{city}, {state}"


def build_messages(schema: Dict[str, Any], user_input: str) -> Iterable[Dict[str, str]]:
    instructions = schema_to_instruction_block(schema)
    system_prompt = (
        "You extract structured real-estate underwriting data from free text.\n"
        "Return ONLY valid JSON with exactly the schema keys.\n"
        "If a required field is missing, return null for that field.\n"
        "Do not add commentary.\n"
        "Informal listing language: phrases such as 'asking $675k', 'listed at 1.2M', "
        "'priced at $450,000', or '$800k OBO' express the asking/list price—map them to "
        "AskingPrice as a JSON number in dollars (e.g. 675000 for $675k). "
        "Do not use numbers from listing dates (e.g. 4/16/2026) as AskingPrice. "
        "RentCast / rent estimate lines (e.g. 'RentCast rent estimate is $4070 per month total') "
        "map to RentCastRent as the monthly dollar integer for the whole property. "
        "Actual occupied rent (e.g. 'in-place rent $5200', 'current collected rent') maps to InPlaceRent—not "
        "third-party estimates. "
        "PurchasePrice is the contract purchase amount; SellerConcessions is seller-paid closing help in dollars. "
        "CityState must be like 'Cleveland, OH' when present. "
        "Required fields must be filled when the text clearly states them."
    )
    user_prompt = (
        "Extract the following fields and normalize values where appropriate.\n"
        f"{instructions}\n\n"
        "Return JSON object with keys exactly as listed above.\n\n"
        f"Input text:\n{user_input}"
    )
    return [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
    ]


def extract_json_from_response(text: str) -> Dict[str, Any]:
    text = text.strip()
    if text.startswith("```"):
        text = text.strip("`")
        if "\n" in text:
            text = text.split("\n", 1)[1]
    start = text.find("{")
    end = text.rfind("}")
    if start == -1 or end == -1 or end <= start:
        raise ValueError(f"Could not locate JSON object in model output: {text}")
    return json.loads(text[start : end + 1])


def _parse_usd_shorthand_to_int(number_part: str, suffix: str | None) -> int:
    n = float(number_part.replace(",", ""))
    suf = (suffix or "").lower()
    if suf == "k":
        n *= 1000
    elif suf == "m":
        n *= 1_000_000
    return int(math.ceil(n))


def _currency_capture_is_date_month(s: str, m: re.Match, num_group: int = 1) -> bool:
    """True if the captured number is the month in an m/d/... date (e.g. listed 4/16/2026)."""
    end = m.end(num_group)
    return end < len(s) and s[end] == "/"


def extract_asking_price_usd_from_text(text: str) -> int | None:
    """
    Best-effort extraction when the model omits or mis-parses AskingPrice.

    Avoids matching bare "listed" before a calendar date (listed 4/16/2026 was misread as $4).
    """
    if not text or not str(text).strip():
        return None
    s = normalize_extraction_source_text(str(text))
    # Order matters: specific anchors first; never bare "listed" without at/for before a number.
    patterns = [
        r"(?is)\basking\b\s*[:\s]*(?:\$|approx\.?\s*)?\s*(\d[\d,]*(?:\.\d+)?)\s*([km])?\b",
        r"(?is)\b(?:listed|priced)\s+(?:at|for)\s+(?:\$|approx\.?\s*)?\s*(\d[\d,]*(?:\.\d+)?)\s*([km])?\b",
        r"(?is)\blist\s+price\b\s*[:\s]*(?:\$|approx\.?\s*)?\s*(\d[\d,]*(?:\.\d+)?)\s*([km])?\b",
        r"(?is)\bfor\s+sale\s+at\b\s+(?:\$|approx\.?\s*)?\s*(\d[\d,]*(?:\.\d+)?)\s*([km])?\b",
        r"(?i)\$\s*(\d[\d,]*(?:\.\d+)?)\s*([km])\b",
        r"(?i)\b(\d[\d,]*(?:\.\d+)?)\s*([km])\b",
    ]
    for pattern in patterns:
        for m in re.finditer(pattern, s):
            if _currency_capture_is_date_month(s, m):
                continue
            num = m.group(1)
            suf = m.group(2) if m.lastindex and m.lastindex >= 2 else None
            try:
                return _parse_usd_shorthand_to_int(num, suf)
            except (TypeError, ValueError):
                continue
    return None


def _copy_field_if_missing(parsed: Dict[str, Any], key: str, value: Any) -> Dict[str, Any]:
    """Return a shallow copy with key set to value only when value is not None and key is absent/null."""
    if value is None or parsed.get(key) is not None:
        return parsed
    merged = dict(parsed)
    merged[key] = value
    return merged


def fill_missing_asking_price_from_source(
    parsed: Dict[str, Any], schema: Dict[str, Any], user_input: str
) -> Dict[str, Any]:
    """If AskingPrice is still null after the LLM, infer it from the raw input text when possible."""
    spec = schema.get("fields", {}).get("AskingPrice")
    if not isinstance(spec, dict) or not spec.get("required"):
        return parsed
    return _copy_field_if_missing(parsed, "AskingPrice", extract_asking_price_usd_from_text(user_input))


def _replace_asking_price_if_bad_model(cur: int | None, hint: int) -> bool:
    """True when regex hint should override the model value."""
    if cur is None:
        return True
    if cur == hint:
        return False
    if hint >= 10000 and cur < 10000:
        return True
    if hint >= 1000 and cur < hint // 100:
        return True
    return False


def reconcile_asking_price_from_source(
    parsed: Dict[str, Any], schema: Dict[str, Any], user_input: str
) -> Dict[str, Any]:
    """
    Replace absurd AskingPrice values from the model (e.g. "4" from a listing date) when the
    source text contains a clear asking/list price (e.g. asking $675k).
    """
    spec = schema.get("fields", {}).get("AskingPrice")
    if not isinstance(spec, dict) or not spec.get("required"):
        return parsed
    hint = extract_asking_price_usd_from_text(user_input)
    if hint is None:
        return parsed

    raw = parsed.get("AskingPrice")
    try:
        cur = None if raw is None else normalize_currency_text_to_int(raw)
    except (TypeError, ValueError):
        cur = None

    if not _replace_asking_price_if_bad_model(cur, hint):
        return parsed
    merged = dict(parsed)
    merged["AskingPrice"] = hint
    return merged


def _first_currency_amount_from_patterns(normalized_text: str, patterns: list[str]) -> int | None:
    """First capturing group in each pattern must be the dollar digits."""
    for pattern in patterns:
        m = re.search(pattern, normalized_text)
        if not m:
            continue
        num = m.group(1).replace(",", "")
        try:
            return normalize_currency_text_to_int(num)
        except (TypeError, ValueError):
            continue
    return None


def extract_rentcast_monthly_total_usd_from_text(text: str) -> int | None:
    """Extract total monthly rent when phrased as RentCast estimate or 'per month'."""
    if not text or not str(text).strip():
        return None
    s = normalize_extraction_source_text(str(text))
    patterns = [
        r"(?is)RentCast\s+rent\s+estimate\s+is\s+\$?\s*(\d[\d,]*)",
        r"(?is)\brent\s+estimate\s+is\s+\$?\s*(\d[\d,]*)",
        r"(?is)\bestimated\s+(?:monthly\s+)?rent\s+is\s+\$?\s*(\d[\d,]*)",
        r"(?is)\$?\s*(\d[\d,]*)\s+per\s+month\s+total\b",
        r"(?is)\$?\s*(\d[\d,]*)\s+per\s+month\b",
    ]
    return _first_currency_amount_from_patterns(s, patterns)


def extract_inplace_rent_monthly_total_usd_from_text(text: str) -> int | None:
    """Extract actual in-place / collected monthly rent (occupied property), distinct from RentCast estimates."""
    if not text or not str(text).strip():
        return None
    s = normalize_extraction_source_text(str(text))
    patterns = [
        r"(?is)\bin[- ]?place\s+rent\b\s*(?:is|=|:)?\s*\$?\s*(\d[\d,]*)",
        r"(?is)\bcurrent\s+(?:collected\s+)?rent\b\s*(?:is|=|:)?\s*\$?\s*(\d[\d,]*)",
        r"(?is)\bactual\s+(?:monthly\s+)?rent\b\s*(?:is|=|:)?\s*\$?\s*(\d[\d,]*)",
        r"(?is)\boccupied\s+rent\b\s*(?:is|=|:)?\s*\$?\s*(\d[\d,]*)",
    ]
    return _first_currency_amount_from_patterns(s, patterns)


def fill_missing_rentcast_rent_from_source(
    parsed: Dict[str, Any], schema: Dict[str, Any], user_input: str
) -> Dict[str, Any]:
    if not isinstance(schema.get("fields", {}).get("RentCastRent"), dict):
        return parsed
    return _copy_field_if_missing(
        parsed, "RentCastRent", extract_rentcast_monthly_total_usd_from_text(user_input)
    )


def fill_missing_inplace_rent_from_source(
    parsed: Dict[str, Any], schema: Dict[str, Any], user_input: str
) -> Dict[str, Any]:
    if not isinstance(schema.get("fields", {}).get("InPlaceRent"), dict):
        return parsed
    return _copy_field_if_missing(
        parsed, "InPlaceRent", extract_inplace_rent_monthly_total_usd_from_text(user_input)
    )


def _enforce_schema_bounds(
    field_name: str, field_spec: Dict[str, Any], value: float, *, as_int: bool
) -> None:
    lo = field_spec.get("min")
    hi = field_spec.get("max")
    if lo is not None:
        bound = int(lo) if as_int else float(lo)
        if value < bound:
            raise ValueError(f"Field '{field_name}' below min {lo}: {value}")
    if hi is not None:
        bound = int(hi) if as_int else float(hi)
        if value > bound:
            raise ValueError(f"Field '{field_name}' above max {hi}: {value}")


def coerce_and_validate(parsed: Dict[str, Any], schema: Dict[str, Any]) -> Dict[str, Any]:
    result: Dict[str, Any] = {}
    for field_name, field_spec in schema["fields"].items():
        raw_value = parsed.get(field_name)
        field_type = field_spec.get("type", "string")

        if raw_value is None:
            if field_spec.get("required", False):
                raise ValueError(f"Required field '{field_name}' is missing/null.")
            result[field_name] = None
            continue

        if field_type == "integer":
            try:
                if normalization_rule_is_currency(field_spec):
                    value = normalize_currency_text_to_int(raw_value)
                else:
                    value = parse_integer_text(raw_value)
            except (TypeError, ValueError) as e:
                raise ValueError(f"Field '{field_name}' must be an integer. Got: {raw_value!r}") from e
            _enforce_schema_bounds(field_name, field_spec, float(value), as_int=True)
            result[field_name] = value
        elif field_type == "number":
            dp = int(field_spec.get("decimal_places", 2))
            if dp < 0:
                raise ValueError(f"Field '{field_name}' has invalid decimal_places: {field_spec.get('decimal_places')}")
            try:
                value = parse_number_text(raw_value, dp)
            except (TypeError, ValueError) as e:
                raise ValueError(f"Field '{field_name}' must be a number. Got: {raw_value!r}") from e
            _enforce_schema_bounds(field_name, field_spec, value, as_int=False)
            result[field_name] = value
        else:
            text = str(raw_value).strip()
            fmt = field_spec.get("format")
            if fmt in ("m/d/yyyy", "mm/dd/yyyy"):
                text = normalize_listing_date_m_d_yyyy(text)
            elif fmt == "City, State":
                text = normalize_city_state_string(text)
            result[field_name] = text

    return result


def run_extraction(
    model_path: Path,
    schema: Dict[str, Any],
    user_input: str,
    n_ctx: int = 4096,
    n_gpu_layers: int = 0,
) -> Dict[str, Any]:
    llm = Llama(
        model_path=str(model_path),
        n_ctx=n_ctx,
        n_gpu_layers=n_gpu_layers,
        verbose=False,
    )
    clean_input = normalize_extraction_source_text(user_input)
    messages = list(build_messages(schema, clean_input))
    completion = llm.create_chat_completion(
        messages=messages,
        temperature=0.0,
        top_p=0.95,
        response_format={"type": "json_object"},
    )
    content = completion["choices"][0]["message"]["content"]
    parsed = extract_json_from_response(content)
    parsed = fill_missing_asking_price_from_source(parsed, schema, clean_input)
    parsed = reconcile_asking_price_from_source(parsed, schema, clean_input)
    parsed = fill_missing_rentcast_rent_from_source(parsed, schema, clean_input)
    parsed = fill_missing_inplace_rent_from_source(parsed, schema, clean_input)
    return coerce_and_validate(parsed, schema)


def resolve_named_cell(workbook, name: str) -> Tuple[Any, str]:
    defined_name = workbook.defined_names.get(name)
    if defined_name is None:
        raise KeyError(f"Named cell/range '{name}' not found in workbook.")

    destinations = list(defined_name.destinations)
    if not destinations:
        raise ValueError(f"Named cell/range '{name}' has no destinations.")

    sheet_name, coord = destinations[0]
    ws = workbook[sheet_name]
    return ws, coord


def write_named_cells(output_workbook_path: Path, values: Dict[str, Any]) -> None:
    wb = load_workbook(output_workbook_path)
    for field_name, value in values.items():
        if value is None:
            continue
        ws, coord = resolve_named_cell(wb, field_name)
        ws[coord] = value
    wb.save(output_workbook_path)


def ensure_output_copy(template_path: Path, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)


_WIN_INVALID_FILENAME = set('<>:"/\\|?*')


def sanitize_filename_component(text: str) -> str:
    """Make a string safe for use as a single Windows filename segment."""
    out = "".join("_" if c in _WIN_INVALID_FILENAME else c for c in text.strip())
    out = re.sub(r"\s+", " ", out).strip(" .")
    return out or "Unknown_Address"


def location_segment_for_output_filename(city_state: str | None) -> str:
    """
    Middle segment of the default output file name: 'Charlotte NC' when CityState is absent,
    otherwise city and state with commas removed (e.g. 'Cleveland, OH' -> 'Cleveland OH').
    """
    if city_state is None:
        return "Charlotte NC"
    s = str(city_state).strip()
    if not s:
        return "Charlotte NC"
    without_commas = s.replace(",", " ")
    collapsed = re.sub(r"\s+", " ", without_commas).strip()
    safe = sanitize_filename_component(collapsed)
    return safe if safe else "Charlotte NC"


def resolve_output_workbook_path(
    output_arg: Path,
    street_address: str,
    today: date,
    city_state: str | None = None,
) -> Path:
    """
    If output_arg ends with .xlsx, use it as the exact file path.
    Otherwise treat output_arg as the output directory and build:
    ``{street} - {location} - Model {yyyy-mm-dd}.xlsx``
    where ``location`` is derived from ``CityState`` when set (comma omitted), else ``Charlotte NC``.
    """
    if output_arg.suffix.lower() == ".xlsx":
        return output_arg
    date_str = today.isoformat()
    x = sanitize_filename_component(street_address)
    loc = location_segment_for_output_filename(city_state)
    filename = f"{x} - {loc} - Model {date_str}.xlsx"
    return output_arg / filename


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract underwriting fields with llama.cpp and write to named Excel cells."
    )
    parser.add_argument(
        "--model",
        default=os.environ.get("UNDERWRITING_GGUF"),
        help=(
            "Path to a GGUF model file. "
            "If omitted, uses the UNDERWRITING_GGUF environment variable."
        ),
    )
    parser.add_argument(
        "--template",
        default="TCG Blank Model Template 2026-04-26.xlsx",
        help="Path to source template workbook.",
    )
    parser.add_argument(
        "--output",
        default="Output",
        help=(
            "Path to the output .xlsx file, or a directory. "
            "If a directory (default: Output), the file name is "
            '"{StreetAddress} - {location} - Model {yyyy-mm-dd}.xlsx" '
            "(location from CityState when present—comma omitted; else Charlotte NC)."
        ),
    )
    parser.add_argument(
        "--schema",
        default="field_schema.yaml",
        help="Path to YAML schema defining extraction fields.",
    )
    parser.add_argument(
        "--prompt",
        default=None,
        help="Natural-language listing text. On Windows PowerShell, use single quotes around "
        "the value if it contains $ (otherwise $675k may be mangled). See also --prompt-file.",
    )
    parser.add_argument(
        "--prompt-file",
        type=Path,
        default=None,
        help="Path to a UTF-8 text file with the listing (avoids shell stripping $ from amounts).",
    )
    parser.add_argument(
        "--gui",
        action="store_true",
        help="Open a window to paste listing text (recommended on Windows; preserves $ in amounts).",
    )
    parser.add_argument("--n-ctx", type=int, default=4096, help="Context window.")
    parser.add_argument(
        "--n-gpu-layers",
        type=int,
        default=0,
        help="Number of model layers to offload to GPU (0 for CPU-only).",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.model:
        raise ValueError(
            "No GGUF model path: pass --model or set the UNDERWRITING_GGUF environment variable."
        )
    model_path = Path(args.model)
    template_path = Path(args.template)
    output_arg = Path(args.output)
    schema_path = Path(args.schema)

    if not model_path.exists():
        raise FileNotFoundError(f"Model file not found: {model_path}")
    if not template_path.exists():
        raise FileNotFoundError(f"Template workbook not found: {template_path}")
    if not schema_path.exists():
        raise FileNotFoundError(f"Schema file not found: {schema_path}")

    if args.gui and (args.prompt is not None or args.prompt_file is not None):
        raise ValueError("Use --gui without --prompt or --prompt-file.")
    if args.prompt_file is not None and args.prompt is not None:
        raise ValueError("Use only one of --prompt and --prompt-file.")

    if args.gui:
        prompt_text = read_prompt_via_gui()
        if not prompt_text:
            raise ValueError("No listing text entered (GUI cancelled or empty).")
    elif args.prompt_file is not None:
        pfile = Path(args.prompt_file)
        if not pfile.is_file():
            raise FileNotFoundError(f"Prompt file not found: {pfile}")
        prompt_text = pfile.read_text(encoding="utf-8")
    elif args.prompt is not None:
        prompt_text = args.prompt
    else:
        raise ValueError("Provide --prompt, --prompt-file, or --gui with the listing text.")

    schema = load_schema(schema_path)
    structured_values = run_extraction(
        model_path=model_path,
        schema=schema,
        user_input=prompt_text,
        n_ctx=args.n_ctx,
        n_gpu_layers=args.n_gpu_layers,
    )

    street = structured_values.get("StreetAddress")
    if not street:
        raise ValueError("StreetAddress is required to build the output filename.")

    output_path = resolve_output_workbook_path(
        output_arg,
        str(street),
        date.today(),
        structured_values.get("CityState"),
    )

    ensure_output_copy(template_path, output_path)
    write_named_cells(output_path, structured_values)

    print("Extracted fields:")
    print(json.dumps(structured_values, indent=2))
    print(f"\nWorkbook written: {output_path}")


if __name__ == "__main__":
    main()
